import openpyxl
from openpyxl import load_workbook
import time
start_time = time.time()
# your code


# data_only=True로 해줘야 수식이 아닌 값으로 받아온다.
filepath="./221116_22년 10월 생손보 월초비례.xlsx"
load_wb = load_workbook(filepath, data_only=True)


# # 셀 주소로 값 출력
# print(load_ws['B2'].value)
elapsed_time = time.time() - start_time
print("Done : file open (",end="")
print(round(elapsed_time,2),end=")")

start_time = time.time()
sheetname='신계약'
# 시트 이름으로 불러오기
load_sheet = load_wb[sheetname]
elapsed_time = time.time() - start_time

def getval(cells):
    result=[]
    Nonecheck=False
    for i in cells:
        if i.value is not None:
            Nonecheck=True
        if Nonecheck:
            result.append(i.value)
        else:
            pass
    if result==[]:
        return None
    return result


namespace = load_sheet['3']
cd={}
idx=0
for i in getval(namespace):
    cd[i]=idx
    idx+=1
print(cd)
get_cells = load_sheet.rows


alltable=[]
for row in get_cells:
    if getval(row) is not None:
        alltable.append(getval(row))

        

#일단 대상보험사 다불러와

goods={ "KDB"    : ["오행복","버팀목"],
        "DGB"    : ["마이솔","그랑","마음든든"],
        "하나"   : ["하나로"],
        "푸르"   : ["100세","달러평생","달러 평생","함께"],
        "삼성"   : ["신성장"],
        "라이나" : ["종신"],
        "메트"   : ["모두의","백만인"],
        "미래"   : ["선택하는"],
        "DB"     : ["알차고","암종신"]
        }

resulttable=[]

for row in alltable[2:]:
    for i in goods:
        for k in goods[i]:
            if i in row[cd["보험사명"]] and k in row[cd["상품명"]] and row[cd["납입주기"]]!="일시납" and row[cd["계약구분"]]!="본인":
                resulttable.append(row)

idx=0
#resulttable = [row for row in resulttable if int(row[cd["납입기간"]])>10]

for row in resulttable:
    if "푸르" in row[cd["보험사명"]] and ("100세" in row[cd["상품명"]] or "달러평생" in row[cd["상품명"]]):
        row.append(int(row[cd["월납화보험료"]]*0.7))
        
    elif int(row[cd["납입기간"]])<10 and "삼성" in row[cd["보험사명"]]:
        row.append(int(row[cd["월납화보험료"]]*0.5))   
        
    elif "기본형" in row[cd["상품명"]] and "라이나" in row[cd["보험사명"]]:
        row.append(0) 
    
    elif "표준형" in row[cd["상품명"]] and "KDB" in row[cd["보험사명"]]:
        row.append(int(row[cd["월납화보험료"]]*0.5)) 
    
    elif "모두의" in row[cd["상품명"]] and "메트" in row[cd["보험사명"]]:
        row.append(int(row[cd["월납화보험료"]]*0.8))    
                   
    elif "표준형" in row[cd["상품명"]] and "푸르" in row[cd["보험사명"]]:
        row.append(int(row[cd["월납화보험료"]]*0))    
        
    elif "함께" in row[cd["상품명"]] and "푸르" in row[cd["보험사명"]] and int(row[cd["납입기간"]])!=5:
        row.append(int(row[cd["월납화보험료"]]*0.7))
                   
    else:
        row.append(int(row[cd["월납화보험료"]]))
        
load_wb.create_sheet('result')
rs=load_wb['result']

for row in resulttable:
    rs.append(row)
load_wb.save("./result11.xlsx")


##
##and조건 => 순차적으로
##or조건 => 중복을 세지않고 더하기
#나중에 모든 column 길이 맞는지 체크해야됨
    
#조건 : 포함 and 나 or / 제외
